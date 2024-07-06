import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl as px
from openpyxl.utils.dataframe import dataframe_to_rows

# Загрузка файла Excel
file_path = r'TestTask2.xlsx'

# Настройки для Chrome
options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')

# Путь к драйверу Chrome
driver_path = r'chromedriver.exe'

# Открываем браузер
driver = webdriver.Chrome(options=options)

# Переход на сайт Google
driver.get('https://www.google.com')

# Словарь для хранения тем и ссылок
theme_links = {}

# Загрузка существующего файла и очистка таблицы
wb = px.load_workbook(file_path)
if 'Sheet1' in wb.sheetnames:
    ws = wb['Sheet1']
    wb.remove(ws)

# Создание нового листа
ws = wb.create_sheet('Sheet1')
ws.append(['Theme', 'Sources'])  # Запись заголовков столбцов

# Загрузка данных из файла Excel в DataFrame
df = pd.read_excel(file_path, sheet_name='Sheet1')

# Собрать все темы в список
themes = df['Theme'].dropna().tolist()

for theme in themes:
    # Явное ожидание загрузки поискового поля
    try:
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
    except Exception as e:
        print(f"Ошибка при ожидании поискового поля: {e}")
        continue

    search_box.clear()
    search_box.send_keys(theme)
    search_box.submit()

    # Явное ожидание загрузки результатов поиска
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.g a'))
        )
    except Exception as e:
        print(f"Ошибка при ожидании результатов поиска для темы '{theme}': {e}")
        theme_links[theme] = []
        continue

    time.sleep(2)  # небольшая пауза для гарантии загрузки

    # Получение первых трех ссылок из результатов поиска
    try:
        links = driver.find_elements(By.CSS_SELECTOR, 'div.g a')[:3]
        theme_links[theme] = [link.get_attribute('href') for link in links]
    except Exception as e:
        print(f"Ошибка при извлечении ссылок для темы '{theme}': {e}")
        theme_links[theme] = []

    time.sleep(2)  # пауза перед следующим запросом

driver.quit()

# Запись данных в новый лист
for theme, links in theme_links.items():
    for link in links:
        ws.append([theme, link])

# Включение фильтрации данных
ws.auto_filter.ref = ws.dimensions

# Сохранение изменений
wb.save(file_path)
