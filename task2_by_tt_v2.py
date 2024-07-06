import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

from send_email import send


def load_excel(file_path, sheet_name='Sheet1'):
    """Загрузка файла Excel и возврат DataFrame."""
    return pd.read_excel(file_path, sheet_name=sheet_name)


def get_themes_from_df(df):
    """Получение списка тем из DataFrame."""
    return df['Theme'].dropna().tolist()


def init_webdriver(driver_path):
    """Инициализация WebDriver с настройками Chrome."""
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    return webdriver.Chrome(options=options)


def search_themes(driver, themes):
    """Поиск тем в Google и возврат словаря с темами и ссылками."""
    theme_links = {}
    driver.get('https://www.google.com')

    for theme in themes:
        try:
            search_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "q"))
            )
            search_box.clear()
            search_box.send_keys(theme)
            search_box.submit()

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.g a'))
            )
            time.sleep(2)

            links = driver.find_elements(By.CSS_SELECTOR, 'div.g a')[:3]
            theme_links[theme] = [link.get_attribute('href') for link in links]

        except Exception as e:
            print(f"Ошибка при обработке темы '{theme}': {e}")
            theme_links[theme] = []

        time.sleep(2)

    return theme_links


def create_new_dataframe(theme_links):
    """Создание нового DataFrame для хранения ссылок."""
    new_rows = [{'Theme': theme, 'Sources': link} for theme, links in theme_links.items() for link in links]
    return pd.DataFrame(new_rows)


def save_to_excel(file_path, df_combined):
    """Сохранение объединенного DataFrame в Excel."""
    wb = load_workbook(file_path)
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
    ws = wb.create_sheet('Sheet1')

    for r in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(r)

    ws.auto_filter.ref = ws.dimensions
    wb.save(file_path)


def main():
    file_path = Path('TestTask2.xlsx')
    driver_path = Path('chromedriver.exe')

    df = load_excel(file_path)
    themes = get_themes_from_df(df)

    with init_webdriver(driver_path) as driver:
        theme_links = search_themes(driver, themes)

    new_df = create_new_dataframe(theme_links)
    df_combined = pd.concat([df, new_df], ignore_index=True)

    save_to_excel(file_path, df_combined)
    send()


if __name__ == "__main__":
    main()
